# encode: utf-8

from __future__ import absolute_import, unicode_literals

import csv
import datetime
import logging
import traceback

from celery.exceptions import MaxRetriesExceededError
from celery.utils.log import get_task_logger
from django.db.models.aggregates import Count
from django.contrib.postgres.aggregates import ArrayAgg, JSONBAgg
from reversion.models import Version, Revision
from django.db import connection
from django.conf import settings
from django.db import transaction
from openpyxl import Workbook

from domain.exceptions import (UnavailableFirebaseServiceError,
                               UnrecorverableFirebaseError)
from domain.firebaser import FirebaseCloudMessaging
from domain.mailer import Mail
from domain.models import (
    MedicalPrescription,
    Patient,
    PrescriptionPiece,
    RejectionReason,
    ScheduledExam,
    ScheduledExamPhoneCall
)
from lab_core.celery import app

from domain.utils import get_phone_call_history

from general_utils.utils import create_zendesk_ticket, create_adf_zendesk_tickets

# Logging handler
log = logging.getLogger(__name__)


@app.task(bind=True, queue="domain.tasks.send_push_notification")
def send_push_notification(self, token, title, message, data=None):
    """
    Schedules a push notification.
    :param self:
    :param token:
    :param title:
    :param message:
    :param data:
    :return:
    """
    # Restrict access to Push Notification:
    if settings.APP_ENVIRONMENT not in (settings.PROD, settings.STAGING, settings.LOCAL):
        log.warn("Restricting firebase access to local, staging and production environments only")
        return

    # For exams PN, check if the current scheduled time is the same when notification was queued, if not, do not send it
    print('send_push_notification: {}'.format(message))
    if data:
        scheduled_exam_id = data.get("scheduled_exam_id", None)
        scheduled_time = data.get("scheduled_time", None)
        scheduled_call_time = data.get("scheduled_call_time", None)
        print('scheduled_exam_id: {}'.format(scheduled_exam_id))
        print('scheduled_time: {}'.format(scheduled_time))
        print('scheduled_call_time: {}'.format(scheduled_call_time))
        if scheduled_exam_id:
            try:
                if scheduled_time:
                    scheduled_exam = ScheduledExam.objects.get(pk=scheduled_exam_id)
                    print('scheduled_exam: {}'.format(scheduled_exam))
                    print('scheduled_exam.status: {}'.format(scheduled_exam.status))

                    if scheduled_exam.scheduled_time:
                        print('scheduled_exam.scheduled_time.timestamp: {}'.format(scheduled_exam.scheduled_time.timestamp()))
                        if int(scheduled_exam.scheduled_time.timestamp()) != scheduled_time or \
                                scheduled_exam.status in (ScheduledExam.PATIENT_CANCELED,
                                                          ScheduledExam.PATIENT_CANCELED_BY_CALL,
                                                          ScheduledExam.LAB_RECORD_CANCELED,
                                                          ScheduledExam.EXAM_EXPIRED):
                            return
                if scheduled_call_time:
                    exam_call = ScheduledExamPhoneCall.objects.get(scheduled_exam=scheduled_exam_id)
                    print('exam_call: {}'.format(exam_call))
                    print('exam_call.call_time.timestamp(): {}'.format(exam_call.call_time.timestamp()))
                    print(int(exam_call.call_time.timestamp()) != scheduled_call_time)
                    if int(exam_call.call_time.timestamp()) != scheduled_call_time:
                        return
            except ScheduledExam.DoesNotExist:
                log.warning(
                    "Could not send push notification. ScheduledExam matching query does not exist"
                )
                return

    if not token:
        return

    # Send notification
    firebase = FirebaseCloudMessaging()
    try:
        response = firebase.send_push_notification(token, title, message, data)

        print(response)
        log.info("Push notification has been sent."
                 "Token: {0}. "
                 "Title: {1}. "
                 "Message: {2}. "
                 "Data: {3}. ".format(token, title, message, data))

    # Retry 10 times with 30 seconds countdown
    except UnavailableFirebaseServiceError:
        log.warning("Could not send push notification. "
                    "Retrying in 30 seconds. "
                    "Token: {0}. "
                    "Title: {1}. "
                    "Message: {2}. "
                    "Data: {3}. ".format(token, title, message, data))
        try:
            self.retry(countdown=30, max_retries=10)

        # When max retries exceeded, send it to an error queue
        except MaxRetriesExceededError:
            log.warning("Could not send push notification. "
                        "Max retries exceeded. "
                        "Token: {0}. "
                        "Title: {1}. "
                        "Message: {2}. "
                        "Data: {3}. ".format(token, title, message, data))

            self.apply_async(
                args=[token, title, message, data],
                queue="domain.tasks.send_push_notification.error"
            )

    # Unrecoverable error, just remove it from the queue
    except (ValueError, UnrecorverableFirebaseError) as e:
        log.warning("Could not send push notification. "
                    "Exception: {0}. "
                    "Token: {1}. "
                    "Title: {2}. "
                    "Message: {3}. "
                    "Data: {4}. ".format(str(e), token, title, message, data))


@app.task(queue="domain.tasks.populate_firebase_week_schedulings_endpoint")
def populate_firebase_week_schedulings_endpoint():
    """
    Populates "/week_schedulings" endpoint on Firebase with all today's exam appointments ('Comparecimento' tab).
    This task runs every midnight with celery-beat.

    It generates a data like this:

    {
        "4": [{
            "insurancePlan": "Company 1",
            "name": "Sandro Santiago",
            "firstExamTime": 1494953118,
            "examNames": ["TGP "],
            "laboratoryName": "Angelica 1832",
            "examIds": [78],
            "avatar": "/media/40/user_data/selfie_1494425656.jpg",
            "labFileCodes": {"78": "asdas2"}
        }, {
            "insurancePlan": "Company 1",
            "name": "user name",
            "firstExamTime": 1494947191,
            "examNames": ["Exam 1", "PCTNEIM"],
            "laboratoryName": "Angelica 1832",
            "examIds": [77, 74],
            "avatar": "/media/1/user_data/selfie_1494583051.jpg",
            "labFileCodes": {"77": "asdasd123", "74": "bcvbcvds31"}
        }],
        "72": [{
            "insurancePlan": "Company 1",
            "name": "user name",
            "firstExamTime": 1494953064,
            "examNames": ["RETORCLIIM"],
            "laboratoryName": "Unid. Alta Coleta Domiciliar - Rj",
            "examIds": [73],
            "avatar": "/media/1/user_data/selfie_1494583051.jpg",
            "labFileCodes": {"73": "asdasd123"}
        }]
    }

    Where keys "4" and "72" are laboratory ids.

    :param:
    :return:
    """
    try:
        import pyrebase
        from domain.models import ScheduledExam

        firebase = pyrebase.initialize_app(settings.FB_CONFIG)
        db = firebase.database()

        next_week = datetime.datetime.now() + datetime.timedelta(days=7)

        all_exams = ScheduledExam.objects.filter(
            scheduled_time__lte=next_week.date(),
            status__in=(ScheduledExam.EXAM_TIME_SCHEDULED, ScheduledExam.LAB_RECORD_OPEN)
        )

        labs = []
        patients = []
        for exam in all_exams.all():
            if exam.laboratory and exam.laboratory not in labs:
                labs.append(exam.laboratory)
            if exam.prescription.patient not in patients:
                patients.append(exam.prescription.patient)

        db.child('week_schedulings').remove()

        for lab in labs:
            fb_data = []
            patients_with_exams = {patient: all_exams.filter(laboratory=lab, prescription__patient=patient).all() for
                                   patient in patients}

            for patient, exams in patients_with_exams.items():
                if not exams:
                    continue

                selfie = None
                if bool(patient.selfie_uploadcare) or bool(patient.selfie):
                    try:
                        selfie = patient.selfie_uploadcare.url
                    except (AttributeError, ValueError):
                        selfie = patient.selfie.url

                first_exam_time = int(exams[0].scheduled_time.timestamp()) if exams[0].scheduled_time else None
                insurance_plan = exams[0].prescription.insurance_company.name if exams[0].prescription.insurance_company else settings.NOT_FOUND

                fb_data.append(
                    {
                        "avatar": selfie,
                        "name": patient.full_name,
                        "laboratoryName": lab.description,
                        "insurancePlan": insurance_plan,
                        "examNames": [scheduled_exam.exam.description for scheduled_exam in exams],
                        "examIds": [scheduled_exam.id for scheduled_exam in exams],
                        "examStatuses": [scheduled_exam.status for scheduled_exam in exams],
                        "firstExamTime": first_exam_time,
                        "labFileCodes": {str(scheduled_exam.id): scheduled_exam.lab_file_code for scheduled_exam in
                                         exams},
                        "scheduledTimes": [
                            int(scheduled_exam.scheduled_time.timestamp()) if scheduled_exam.scheduled_time else None
                            for scheduled_exam in exams]
                    })

            db.child('week_schedulings').child(lab.id).set(fb_data)

    except Exception as e:
        traceback.print_tb(e.__traceback__)
        log.error("Unable to sync with firebase from celery-beat (populate_firebase_week_schedulings_endpoint) {}"
                  .format(e))


@app.task(queue="domain.tasks.populate_firebase_results_ac_endpoint")
def populate_firebase_results_ac_endpoint():
    """
    Populates "/results_ac" endpoint on Firebase with all exam results ('Resultados AC' tab).
    This task runs every midnight with celery-beat and on every exam status change regarding results..

    It generates a data like this:

    {
        lab_id_1: {								//time_stamp
            exam_id_1:{
                avatar: String,
                patient_name: String,
                insurance_plan: String,
                exam_name: String,
                exam_id: int,
                patient_id: int,
                prescription_id: int,
                result_expected_at: timestamp
            },
            exam_id_2:{
                avatar: String,
                patient_name: String,
                insurance_plan: String,
                exam_name: String,
                exam_id: int,
                patient_id: int,
                prescription_id: int,
                result_expected_at: timestamp
            }
        },
        lab_id_2: {								//time_stamp
            exam_id_3:{
                avatar: String,
                patient_name: String,
                insurance_plan: String,
                exam_name: String,
                exam_id: int,
                patient_id: int,
                prescription_id: int,
                result_expected_at: timestamp
            }
        }
    }

    :param:
    :return:
    """
    try:
        import pyrebase
        from domain.models import ScheduledExam, Exam

        firebase = pyrebase.initialize_app(settings.FB_CONFIG)
        db = firebase.database()

        pending_result_exams = ScheduledExam.objects.filter(
            exam__exam_type=Exam.AC,
            status__in=(ScheduledExam.PROCEDURES_EXECUTED, ScheduledExam.RESULTS_DELAYED)
        )

        labs = []
        for exam in pending_result_exams.all():
            if exam.laboratory and exam.laboratory not in labs:
                labs.append(exam.laboratory)

        db.child('results_ac').remove()

        for lab in labs:
            fb_data = []
            exams_by_lab = pending_result_exams.filter(laboratory=lab).all()

            for exam in exams_by_lab:

                selfie = None
                if bool(exam.prescription.patient.selfie_uploadcare) or bool(exam.prescription.patient.selfie):
                    try:
                        selfie = exam.prescription.patient.selfie_uploadcare.url
                    except (AttributeError, ValueError):
                        selfie = exam.prescription.patient.selfie.url

                insurance_plan = exam.prescription.insurance_company.name if exam.prescription.insurance_company else settings.NOT_FOUND

                if exam.results_expected_at:
                    result_expected_at = datetime.datetime(year=exam.results_expected_at.year,
                                                           month=exam.results_expected_at.month,
                                                           day=exam.results_expected_at.day
                                                           )
                    result_expected_at = int(result_expected_at.timestamp())
                else:
                    result_expected_at = None

                fb_data.append(
                    {
                        "avatar": selfie,
                        "name": exam.prescription.patient.full_name,
                        "labFileCode": exam.lab_file_code,
                        "insurancePlan": insurance_plan,
                        "examName": exam.exam.name,
                        "examId": exam.id,
                        "patientId": exam.prescription.patient.user.id,
                        "PrescriptionId": exam.prescription.id,
                        "resultExpectedAt": result_expected_at,
                        "labName": lab.description,
                    })
            if fb_data:
                db.child('results_ac').child(str(lab.id)).set(fb_data)

    except Exception as e:
        traceback.print_tb(e.__traceback__)
        log.error(
            "Unable to sync with firebase from celery-beat (populate_firebase_results_ac_endpoint) {}"
            .format(e))


@app.task(queue="domain.tasks.populate_firebase_results_rdi_endpoint")
def populate_firebase_results_rdi_endpoint():
    """
    Populates "/results_rdi" endpoint on Firebase with all exam results. ('Resultados RDI' tab)
    This task runs every midnight with celery-beat and on every exam status change regarding results.

    It generates a data like this:

    {
        lab_id_1: {								//time_stamp
            exam_id_1:{
                avatar: String,
                patient_name: String,
                insurance_plan: String,
                exam_name: String,
                exam_id: int,
                patient_id: int,
                prescription_id: int,
                result_expected_at: timestamp
            },
            exam_id_2:{
                avatar: String,
                patient_name: String,
                insurance_plan: String,
                exam_name: String,
                exam_id: int,
                patient_id: int,
                prescription_id: int,
                result_expected_at: timestamp
            }
        },
        lab_id_2: {								//time_stamp
            exam_id_3:{
                avatar: String,
                patient_name: String,
                insurance_plan: String,
                exam_name: String,
                exam_id: int,
                patient_id: int,
                prescription_id: int,
                result_expected_at: timestamp
            }
        }
    }

    :param:
    :return:
    """
    try:
        import pyrebase
        from domain.models import ScheduledExam, Exam

        firebase = pyrebase.initialize_app(settings.FB_CONFIG)
        db = firebase.database()

        pending_result_exams = ScheduledExam.objects.filter(
            exam__exam_type=Exam.RDI,
            status__in=(ScheduledExam.PROCEDURES_EXECUTED, ScheduledExam.RESULTS_DELAYED)
        )

        labs = []
        for exam in pending_result_exams.all():
            if exam.laboratory and exam.laboratory not in labs:
                labs.append(exam.laboratory)

        db.child('results_rdi').remove()

        for lab in labs:
            fb_data = []
            exams_by_lab = pending_result_exams.filter(laboratory=lab).all()

            for exam in exams_by_lab:

                selfie = None
                if bool(exam.prescription.patient.selfie_uploadcare) or bool(exam.prescription.patient.selfie):
                    try:
                        selfie = exam.prescription.patient.selfie_uploadcare.url
                    except (AttributeError, ValueError):
                        selfie = exam.prescription.patient.selfie.url

                insurance_plan = exam.prescription.insurance_company.name if exam.prescription.insurance_company else settings.NOT_FOUND

                if exam.results_expected_at:
                    result_expected_at = datetime.datetime(year=exam.results_expected_at.year,
                                                           month=exam.results_expected_at.month,
                                                           day=exam.results_expected_at.day
                                                           )
                    result_expected_at = int(result_expected_at.timestamp())
                else:
                    result_expected_at = None

                fb_data.append(
                    {
                        "avatar": selfie,
                        "name": exam.prescription.patient.full_name,
                        "insurancePlan": insurance_plan,
                        "examName": exam.exam.name,
                        "examId": exam.id,
                        "patientId": exam.prescription.patient.user.id,
                        "PrescriptionId": exam.prescription.id,
                        "resultExpectedAt": result_expected_at,
                        "labName": lab.description,
                    })
            if fb_data:
                db.child('results_rdi').child(str(lab.id)).set(fb_data)

    except Exception as e:
        traceback.print_tb(e.__traceback__)
        log.error(
            "Unable to sync with firebase from celery-beat (populate_firebase_results_rdi_endpoint) {}"
            .format(e))


def populate_firebase_calls_endpoint(scheduled_phone_call):
    """
    Populates "/calls" endpoint on Firebase with all exam results. ('Chamadas' tab)

    It generates a data like this:

    {
        prescription_id_1: {

            phone_call_id_1: {
                avatar: String,
                insurancePlan: String,
                examName: [String],
                prescriptionId: String,
                name: int,
                prefferedLabs: [int],
                patientId: int,
                scheduledExamId: [int],
                key: [int],
                callTime: [timestamp]
            }
        },
        prescription_id_2: {

            phone_call_id_2: {
                avatar: String,
                insurancePlan: String,
                examName: [String],
                prescriptionId: String,
                name: int,
                prefferedLabs: [int],
                patientId: int,
                scheduledExamId: [int],
                key: [int],
                callTime: [timestamp]
            }
        }
    }

    :param:
    :return:
    """
    try:
        import pyrebase
        from domain.models import Exam, ScheduledExam, ScheduledExamPhoneCall

        firebase = pyrebase.initialize_app(settings.FB_CONFIG)
        db = firebase.database()

        prescription = scheduled_phone_call.scheduled_exam.prescription
        insurance_company_name = prescription.insurance_company.name if prescription.insurance_company else settings.NOT_FOUND

        selfie = None
        if bool(prescription.patient.selfie_uploadcare) or bool(prescription.patient.selfie):
            try:
                selfie = prescription.patient.selfie_uploadcare.url
            except (AttributeError, ValueError):
                selfie = prescription.patient.selfie.url

        fb_data = {
            "avatar": selfie,
            "insurancePlan": insurance_company_name,
            "examName": scheduled_phone_call.scheduled_exam.exam.name,
            "prescriptionId": prescription.pk,
            "name": prescription.patient.full_name,
            "prefferedLabs": [str(lab) for lab in prescription.patient.preferred_laboratories.all()],
            "patientId": prescription.patient.user.id,
            "scheduledExamId": scheduled_phone_call.scheduled_exam.id,
            "key": scheduled_phone_call.pk,
            "callTime": int(scheduled_phone_call.call_time.timestamp())
        }

        db.child('calls').child(prescription.id).child(scheduled_phone_call.pk).set(fb_data)

    except Exception as e:
        traceback.print_tb(e.__traceback__)
        log.error("Unable to sync with firebase /calls endpoint. {0}".format(e))


@app.task(queue="domain.tasks.expire_prescription_pieces")
@transaction.atomic
def expire_prescription_pieces():
    """
    Sets all prescription pieces status as REQUEST_EXPIRED if expiration_date is older or equals today.
    It runs once a day at midnight via celery-beat.
    :param:
    :return:
    """
    try:
        from domain.models import MedicalPrescription, PrescriptionPiece
        from domain.signals.update_firebase_db import remove_nested_nodes_from_expired_prescription_piece

        expired_pieces = PrescriptionPiece.objects.filter(expiration_date__lte=datetime.date.today())
        expired_pieces = expired_pieces.exclude(status=PrescriptionPiece.REQUEST_EXPIRED)
        with transaction.atomic():
            for piece in expired_pieces:
                piece.status = PrescriptionPiece.REQUEST_EXPIRED
                piece.save()
                rejected_expired_reason = RejectionReason.objects.get(
                    status=RejectionReason.REJECTED_REQUEST_EXPIRED
                )
                piece.rejection_reasons.add(rejected_expired_reason)
                remove_nested_nodes_from_expired_prescription_piece(piece)

                prescription = piece.prescription
                prescription_pieces = prescription.pieces.all()
                not_expired_pieces = prescription_pieces.exclude(status=PrescriptionPiece.REQUEST_EXPIRED).count()

                if prescription.pieces.count() == 1 or (prescription_pieces.count() > 1 and not_expired_pieces == 0):
                    prescription.status = MedicalPrescription.REQUEST_EXPIRED
                    prescription.save()
                    prescription.rejection_reasons.add(rejected_expired_reason)

    except Exception as e:
        traceback.print_tb(e.__traceback__)
        log.error(
            "Unable to expire prescription piece {}"
            .format(e))


@app.task(queue='domain.tasks.zen_desk')
def zendesk_prescription(prescription):
    """
    creates the zendesk ticket as an async action
    :return:
    """
    create_zendesk_ticket(prescription)

@app.task(queue='domain.tasks.zendesk_adf')
def zendesk_adf(exams):
    """
    updates the exams for ADF
    :return:
    """
    print('create_adf_zendesk_tickets scheduled {}'.format(exams))
    create_adf_zendesk_tickets(exams)

@app.task(queue="domain.tasks.send_users_report")
def send_users_report():
    """
    Sends users numbers (since 15th June 2017) report by email everyday at 5pm
    :return:
    """
    # Restricting access:
    if settings.APP_ENVIRONMENT not in (settings.LOCAL, settings.PROD):
        log.warn("Restricting access to local and production environments only")
        return

    try:
        from domain.models import Patient, ScheduledExam, ScheduledExamPhoneCall
        from patient_api.utils import get_scheduled_exam_version

        # initial_date = datetime.date(year=2017, month=6, day=15)
        initial_date = datetime.datetime.now() - datetime.timedelta(days=7)
        sheet_headers = ("email", "phone", "name", "registration_date", "is_confirmed", "has_prescriptions",
                         "scheduled_call", "scheduled_exam", "attended_exam")
        file_name = "users_report.csv"

        # Writes file:
        with open(file_name, 'w+') as csvfile:
            writer = csv.writer(csvfile)

            # Header
            writer.writerow(sheet_headers)

            # Rows
            for patient in Patient.objects.filter(user__date_joined__gte=initial_date).order_by('-user__date_joined'):
                is_confirmed = "yes" if patient.is_confirmed else "no"
                has_prescriptions = "yes" if patient.medicalprescription_set.exists() else "no"
                scheduled_call, scheduled_exam, attended_exam = "no", "no", "no"

                # Iterate over prescriptions/exams:
                for prescription in patient.medicalprescription_set.all():
                    for exam in prescription.scheduledexam_set.all():

                        if ScheduledExamPhoneCall.objects.filter(scheduled_exam=exam.id).exists():
                            scheduled_call = "yes"

                        exam_statuses = get_scheduled_exam_version(exam, "status")
                        for status in exam_statuses:
                            if status["value"] == ScheduledExam.EXAM_TIME_SCHEDULED:
                                scheduled_exam = "yes"
                            if status["value"] == ScheduledExam.PROCEDURES_EXECUTED:
                                attended_exam = "yes"

                row = (patient.user.email, str(patient.phone), patient.full_name, str(patient.user.date_joined.date()),
                       is_confirmed, has_prescriptions, scheduled_call, scheduled_exam, attended_exam)

                print(row)
                writer.writerow(row)

        # Sends file via email:
        text = """
                <html>
                    Hey!
                    <br><br>
                    This is your daily Users report since 15th June, 2017.
                    <br><br>
                    Sara
                    <br><br>
                    --
                    <br><br>
                    This is an automatic message, do not reply.
                </html>
                """
        response = Mail.send(to=settings.USERS_REPORT_LIST,
                             subject="Daily Users Report",
                             text=text,
                             html=True,
                             attachments=file_name)

        print("Email sent: {0}".format(response))

    except Exception as e:
        print(e)
        log.error(
            "Unable to build users report {}"
            .format(e))


@app.task(queue="domain.tasks.send_users_orders_report")
def send_users_orders_report():
    """
    Sends users orders for prescription/exams (since 15th June 2017) report by email everyday at 5pm
    from domain.tasks import send_kpis_per_day_report, send_general_kpi, send_kpi_prescription_call, send_users_orders_report, send_users_report

    :return:
    """
    # Restricting access:
    if settings.APP_ENVIRONMENT not in (settings.LOCAL, settings.PROD):
        log.warn("Restricting access to local and production environments only")
        return

    try:
        from domain.models import Patient, ScheduledExam, MedicalPrescription
        from patient_api.utils import get_scheduled_exam_version, get_prescription_version

        # initial_date = datetime.date(year=2017, month=6, day=15)
        initial_date = datetime.datetime.now() - datetime.timedelta(days=7)
        sheet_headers = ("email", "phone/prescription", "exam", "status", "status_modified_date",
                         "status_modified_time", "URL")
        file_name = "users_orders_report.csv"

        # Writes file:
        with open(file_name, 'w+') as csvfile:
            writer = csv.writer(csvfile)

            # Header
            writer.writerow(sheet_headers)

            # Rows
            for patient in Patient.objects.filter(user__date_joined__gte=initial_date).order_by('-user__date_joined'):
                # Iterate over prescriptions/exams:
                prescriptions = patient.medicalprescription_set.all().order_by('-created')
                if not prescriptions:
                    continue

                # Writes patient:
                row = (patient.user.email, str(patient.phone))
                print(row)
                writer.writerow(row)

                # Writes prescriptions/exams:
                for prescription in prescriptions:
                    exams = prescription.scheduledexam_set.all().order_by('-created')
                    if not exams:
                        prescription_statuses = get_prescription_version(prescription, "status")
                        prescription_status_date = None

                        for status in prescription_statuses:
                            if prescription.status == MedicalPrescription.REQUEST_EXPIRED:
                                prescription_status_date = prescription.expiration_date or prescription.modified
                                break
                            elif prescription.status == status["value"]:
                                prescription_status_date = datetime.datetime.fromtimestamp(status["modified"])

                        row = ("", "prescription {}".format(prescription.pk), "No exams", prescription.status,
                               prescription_status_date.date() if prescription_status_date else None,
                               prescription_status_date.time() if prescription_status_date else None,
                               "{0}/admin/domain/medicalprescription/{1}/".format(settings.DOMAIN_NAME, prescription.pk))
                        print(row)
                        writer.writerow(row)
                        continue

                    row = ("", "prescription {0}".format(prescription.pk))
                    print(row)
                    writer.writerow(row)

                    for exam in exams:
                        exam_statuses = get_scheduled_exam_version(exam, "status")
                        exam_status_date = None

                        for status in exam_statuses:
                            if exam.status == ScheduledExam.EXAM_EXPIRED:
                                exam_status_date = exam.prescription.expiration_date or exam.prescription.modified
                                break
                            elif exam.status == status["value"]:
                                exam_status_date = datetime.datetime.fromtimestamp(status["modified"])

                        row = ("", "", exam.exam.name, exam.status,
                               exam_status_date.date() if exam_status_date else None,
                               exam_status_date.time() if exam_status_date else None,
                               "{0}/admin/domain/scheduledexam/{1}/".format(settings.DOMAIN_NAME, exam.pk))
                        print(row)
                        writer.writerow(row)

        # Sends file via email:
        text = """
                <html>
                    Hi!
                    <br><br>
                    This is your daily Users orders report since {}
                    <br><br>
                    Sara
                    <br><br>
                    --
                    <br><br>
                    This is an automatic message, do not reply.
                </html>
                """.format(initial_date.isoformat())
        response = Mail.send(to=settings.USERS_REPORT_LIST,
                             subject="Daily Users Orders Report",
                             text=text,
                             html=True,
                             attachments=file_name)

        print("Email sent: {0}".format(response))

    except Exception as e:
        print(e)
        log.error(
            "Unable to build users report {}"
            .format(e))




@app.task(queue="domain.tasks.send_kpi_prescription_call")
def send_kpi_prescription_call():
    """
    Sends kpis report (since 15th June 2017) by email everyday at 5pm
    :return:
    """
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo

    file_name = 'Atendimento_Sara_NAC_{}.xls'.format(datetime.date.today().isoformat())
    wb = Workbook()
    writer1 = wb.active
    writer1.title = datetime.date.today().isoformat()
    writer1.append(('Código',
                    'Paciente',
                    'E-mail',
                    'Recebimento',
                    'Análise',
                    'Ligação 1',
                    'Tentativa 1',
                    'Ligação 2',
                    'Tentativa 2',
                    'Ligação 3',
                    'Tentativa 3',
                    'Ligação 4',
                    'Tentativa 4',
                    'Ligação 5',
                    'Tentativa 5'))

    for i in MedicalPrescription.objects.filter(status=MedicalPrescription.PATIENT_REQUESTED).order_by('created'):
        writer1.append((i.id,
                        i.patient.full_name,
                        i.patient.user.email,
                        i.created))

    iids = []
    for i in ScheduledExam.objects.filter(status=ScheduledExam.PHONE_CALL_SCHEDULED).order_by('created'):
        if i.prescription_id not in iids:
            history = get_phone_call_history(i)

            row = [i.prescription_id,
                   i.prescription.patient.full_name,
                   i.prescription.patient.user.email,
                   i.prescription.created,
                   i.created]

            for h in history:
                row.append(h.get('scheduled_time'))
                row.append(h.get('call_time'))

            writer1.append(row)

            iids.append(i.prescription_id)

    tab = Table(displayName="PrescricaoLigacao", ref="A1:O{}".format(writer1.max_row))

    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    #
    # wb.add_named_style(NamedStyle(name='custom_datetime', number_format='YYYY-MM-DD HH:MM:MM'))
    #
    # for cell in writer1['D2':'O{}'.format(writer1.max_row)]:
    #     for c in cell:
    #         c.style = 'custom_datetime'

    tab.tableStyleInfo = style
    writer1.add_table(tab)
    wb.save(file_name)

    text = u"""
                <html>
                    Relatório de atendimento Prescrição e Ligação
                    <br><br>
                    Data: {}
                    <br><br>
                    Sara
                    <br><br>
                </html>
                """.format(datetime.date.today().isoformat())
    response = Mail.send(to=['sandro.lourenco@roundpe.gs'],
                         subject="Relatório de atendimento Prescrição e Ligação",
                         text=text,
                         html=True,
                         attachments=[file_name, ])
    print("Email sent: {0}".format(response))

@app.task(queue="domain.tasks.send_general_kpi")
def send_general_kpi():
    from django.template.loader import render_to_string
    from collections import OrderedDict

    prescriptionsbyday = MedicalPrescription.objects.filter(created__gt='2017-09-01').extra(
        {"day": "date_trunc('day', domain_medicalprescription.created)"}) \
        .values("day").order_by().annotate(count=Count("id"))
    prescriptionsbymonth = MedicalPrescription.objects.filter(created__gt='2017-09-01').extra(
        {"month": "date_trunc('month', domain_medicalprescription.created)"}) \
        .values("month").order_by().annotate(count=Count("id"))

    ss = (ScheduledExam.PROCEDURES_EXECUTED, ScheduledExam.RESULTS_DELAYED,
          ScheduledExam.RESULTS_RECEIVED, ScheduledExam.LAB_RECORD_OPEN)

    scbymonth = ScheduledExam.objects.filter(created__gt='2017-09-01', status__in=ss).extra(
        {"month": "date_trunc('month', domain_scheduledexam.scheduled_time)"}) \
        .values("month", "prescription_id").order_by().annotate(count=Count("prescription_id"))

    scbyday = ScheduledExam.objects.filter(created__gt='2017-09-01', status__in=ss).extra(
        {"day": "date_trunc('day', domain_scheduledexam.scheduled_time)"}) \
        .values("day", "prescription_id").order_by().annotate(count=Count("prescription_id"))

    byday = {}
    bymonth = {}
    totalmonth = {'schedule': 0,'prescription': 0}

    for i in prescriptionsbyday:
        byday[i.get('day').strftime('%Y-%m-%d')] = {'prescription': int(i.get('count', 0)) + byday.get(i.get('day').strftime('%Y-%m-%d'), {}).get('prescription', 0)}

    for i in scbyday:
        try:
            byday.get(i.get('day').strftime('%Y-%m-%d'))['schedule'] = 1 + byday.get(i.get('day').strftime('%Y-%m-%d'), {}).get('schedule')
        except Exception:
            byday[i.get('day').strftime('%Y-%m-%d')]['schedule'] = 1



    for i in prescriptionsbymonth:
        bymonth[i.get('month').strftime('%Y-%m')] = {'prescription': int(i.get('count', 0))}
        print(bymonth)
        totalmonth['prescription'] = totalmonth['prescription'] + int(i.get('count', 0))

    for i in scbymonth:
        try:
            bymonth.get(i.get('month').strftime('%Y-%m'))['schedule'] = 1 + bymonth[i.get('month').strftime('%Y-%m')].get('schedule')
        except Exception:
            bymonth[i.get('month').strftime('%Y-%m')]['schedule'] = 1
        totalmonth['schedule'] = totalmonth['schedule'] + 1

    byday = OrderedDict(sorted(byday.items()))
    bymonth = OrderedDict(sorted(bymonth.items()))

    bymonth['total'] = totalmonth

    rendered = render_to_string('domain/kpi_base.html', {'day': byday, 'month': bymonth})
    cc = settings.USERS_REPORT_LIST + settings.DATA_REPORT_LIST
    # cc = ['sandro.lourenco@roundpe.gs']
    response = Mail.send(to=cc,
                         subject="Relatório de atendimento na unidade",
                         text=rendered,
                         html=True)
    print("Email sent: {0}".format(response))

@app.task(queue="domain.tasks.schedules_status")
def schedules_status():
    from django.template.loader import render_to_string
    from collections import OrderedDict

    prescriptionsbyday = MedicalPrescription.objects.filter(created__gt='2017-09-01').extra(
        {"day": "date_trunc('day', domain_medicalprescription.created)"}) \
        .values("day").order_by().annotate(count=Count("id"))
    prescriptionsbymonth = MedicalPrescription.objects.filter(created__gt='2017-09-01').extra(
        {"month": "date_trunc('month', domain_medicalprescription.created)"}) \
        .values("month").order_by().annotate(count=Count("id"))

    scbymonth = ScheduledExam.objects.filter(created__gt='2017-09-01', status__in=(ScheduledExam.PROCEDURES_EXECUTED, ScheduledExam.RESULTS_DELAYED, ScheduledExam.RESULTS_RECEIVED)).extra(
        {"month": "date_trunc('month', domain_scheduledexam.scheduled_time)"}) \
        .values("month", "prescription_id").order_by().annotate(count=Count("prescription_id"))

    scbyday = ScheduledExam.objects.filter(created__gt='2017-09-01', status__in=(ScheduledExam.PROCEDURES_EXECUTED, ScheduledExam.RESULTS_DELAYED, ScheduledExam.RESULTS_RECEIVED)).extra(
        {"day": "date_trunc('day', domain_scheduledexam.scheduled_time)"}) \
        .values("day", "prescription_id").order_by().annotate(count=Count("prescription_id"))


    byday = {}
    bymonth = {}
    totalmonth = {'schedule': 0,'prescription': 0}

    for i in prescriptionsbyday:
        byday[i.get('day').strftime('%Y-%m-%d')] = {'prescription': int(i.get('count', 0)) + byday.get(i.get('day').strftime('%Y-%m-%d'), {}).get('prescription', 0)}

    for i in scbyday:
        try:
            byday.get(i.get('day').strftime('%Y-%m-%d'))['schedule'] = int(i.get('count', 0)) + byday.get(i.get('day').strftime('%Y-%m-%d'), {}).get('schedule')
        except Exception:
            byday[i.get('day').strftime('%Y-%m-%d')]['schedule'] = int(i.get('count', 0))



    for i in prescriptionsbymonth:
        bymonth[i.get('month').strftime('%Y-%m')] = {'prescription': int(i.get('count', 0))}
        print(bymonth)
        totalmonth['prescription'] = totalmonth['prescription'] + int(i.get('count', 0))

    for i in scbymonth:
        try:
            bymonth.get(i.get('month').strftime('%Y-%m'))['schedule'] = int(i.get('count', 0)) + bymonth[i.get('month').strftime('%Y-%m')].get('schedule')
        except Exception:
            bymonth[i.get('month').strftime('%Y-%m')]['schedule'] = int(i.get('count', 0))
        totalmonth['schedule'] = totalmonth['schedule'] + int(i.get('count', 0))

    byday = OrderedDict(sorted(byday.items()))
    bymonth = OrderedDict(sorted(bymonth.items()))

    bymonth['total'] = totalmonth
    cc = settings.USERS_REPORT_LIST + settings.DATA_REPORT_LIST
    rendered = render_to_string('domain/kpi_base.html', {'day': byday, 'month': bymonth})
    response = Mail.send(to=cc,
                         subject="Relatório de atendimento na unidade",
                         text=rendered,
                         html=True)
    print("Email sent: {0}".format(response))

@app.task(queue="domain.tasks.send_kpis_per_day_report")
def send_kpis_per_day_report():
    """
    Sends kpis report (since 15th June 2017) by email everyday at 5pm
    :return:
    """
    # Restricting access:
    if settings.APP_ENVIRONMENT not in (settings.LOCAL, settings.PROD):
        log.warn("Restricting access to local and production environments only")
        return

    try:
        from domain.models import Patient, ScheduledExam, MedicalPrescription
        from patient_api.utils import get_scheduled_exam_version, get_prescription_version

        today = datetime.date.today()
        numdays = 15
        date_list = [today - datetime.timedelta(days=x) for x in range(0, numdays + 1)]
        all_exams = ScheduledExam.objects.all()

        sheet_headers = ("date", "registrations",
                         "total_prescriptions_created", "users_who_created_prescriptions",
                         "total_calls_created", "users_who_created_calls",
                         "total_exams_created", "users_who_created_exams",
                         "total_exams_created_with_future_scheduling", "users_who_created_exams_with_future_scheduling",
                         "total_exams_created_with_cancelation", "users_who_created_exams_with_cancelation",
                         "total_exams_created_attended_on_future", "users_who_created_exams_attended_on_future",
                         "todays_attended_exams", "users_who_attended_exam",
                         "todays_scheduled_exams", "users_who_scheduled_exam", "users_with_scheduled_exams_for_future")

        file_name = "kpis_per_day_report.xlsx"
        kpis_per_day_report_tab = "kpis_per_day_report"
        users_who_created_prescriptions_file = "users_who_created_prescriptions.csv"
        users_who_created_calls_file = "users_who_created_calls.csv"
        users_who_created_exams_file = "users_who_created_exams.csv"
        users_who_attended_exam_file = "users_who_attended_exam.csv"
        users_who_scheduled_exam_file = "users_who_scheduled_exam.csv"
        users_with_scheduled_exams_for_today_file = "users_with_scheduled_exams_for_today.csv"
        users_with_calls_for_today_file = "users_with_calls_for_today.csv"
        users_with_scheduled_exams_for_future_file = "users_with_scheduled_exams_for_future.csv"

        # Writes file:
        import os
        from openpyxl import Workbook
        wb = Workbook()
        writer1 = wb.active
        writer1.title = kpis_per_day_report_tab
        writer2 = wb.create_sheet(users_who_created_prescriptions_file)
        writer3 = wb.create_sheet(users_who_created_calls_file)
        writer4 = wb.create_sheet(users_who_created_exams_file)
        writer5 = wb.create_sheet(users_who_attended_exam_file)
        writer6 = wb.create_sheet(users_who_scheduled_exam_file)
        writer7 = wb.create_sheet(users_with_scheduled_exams_for_today_file)
        writer8 = wb.create_sheet(users_with_calls_for_today_file)
        writer9 = wb.create_sheet(users_with_scheduled_exams_for_future_file)

        writer1.append(sheet_headers)

        qregistrations = Patient.objects.all().extra({"day": "date_trunc('day', created)"}).values(
            "day").order_by().annotate(count=Count("user_id"))

        registrations = {i.get('day').date(): i.get('count') for i in qregistrations}

        qprescriptions = MedicalPrescription.objects.all().extra(
            {"day": "date_trunc('day', domain_medicalprescription.created)"}) \
            .values("day").order_by().annotate(pids=ArrayAgg('patient__user__email'))

        prescriptionuser = MedicalPrescription.objects.all().values("id", "patient__user__email")

        qexamver = ScheduledExamPhoneCall.objects.all().extra(
            {"day": "date_trunc('day', domain_scheduledexamphonecall.created)"}) \
            .values("day").order_by().annotate(pids=ArrayAgg('scheduled_exam__prescription__patient__user__email'))

        examspc = {}

        with connection.cursor() as cursor:
            cursor.execute("""select date_trunc('day', cast(serialized_data::json->0->'fields'->>'created' as timestamp)) as day, serialized_data::json->0->'fields'->>'status' as status, count(serialized_data::json->0->'fields'->>'status') as total from reversion_version where content_type_id = 30 group by day, status;""")
            for row in cursor.fetchall():
                dd = row[0].date()
                if not examspc.get(dd, None):
                    examspc[dd] = {}
                examspc[dd].update({row[1]: int(row[2]) or 0})

        all_exams = {}

        # with connection.cursor() as cursor:
        #     cursor.execute("""select date_trunc('day', cast(serialized_data::json->0->'fields'->>'modified' as timestamp)) as day, serialized_data::json->0->'fields'->>'status' as status from reversion_version where content_type_id = 33""")
        #     for row in cursor.fetchall():
        #         dd = row[0].date()
        #         if not all_exams.get(dd, None):
        #             all_exams[dd] = {}
        #         all_exams[dd].update({row[1]: all_exams[dd].get(row[1], 0) + 1})

        with connection.cursor() as cursor:
            cursor.execute("""select date_trunc('day', cast(serialized_data::json->0->'fields'->>'created' as timestamp)) as day, serialized_data::json->0->'fields'->>'status' as status, count(serialized_data::json->0->'fields'->>'status') as total from reversion_version where content_type_id = 33 group by day, status;""")
            for row in cursor.fetchall():
                dd = row[0].date()
                if not all_exams.get(dd, None):
                    all_exams[dd] = {}
                all_exams[dd].update({row[1]: int(row[2]) or 0})


        prescriptions = {i.get('day').date(): i.get('pids') for i in qprescriptions}

        qexams = ScheduledExam.objects.all().extra({"day": "date_trunc('day', domain_scheduledexam.created)"}) \
            .values("day", "status").order_by().annotate(pids=ArrayAgg('prescription__patient__user__email'))

        exams = {}

        for ex in qexams:
            dd = ex.get('day').date()
            if not exams.get(dd, None):
                exams[dd] = {}
            exams[dd].update({ex.get('status'): ex.get('pids')})

        # Rows
        for date in date_list:

            # registrations = Patient.objects.filter(user__date_joined__date=date).count()

            cprescriptions = prescriptions.get(date, [])
            users_who_created_prescriptions = set(cprescriptions)
            total_prescriptions_created = len(cprescriptions)

            users_who_created_calls = len(set(examspc.get(date, [])))

            total_calls_created = len(examspc.get(date, []))

            allusers = []
            if exams.get(date):
                for k,v in exams.get(date).items():
                    allusers += v
            users_who_created_exams = set(allusers)

            total_exams_created = len(allusers)
            users_who_created_exams_with_future_scheduling = set()
            users_who_created_exams_with_cancelation = set()
            users_who_created_exams_attended_on_future = set()
            total_exams_created_with_future_scheduling = 0
            total_exams_created_with_cancelation = 0
            total_exams_created_attended_on_future = 0

            exam = exams.get(date)
            if exam:
                for k, v in exam.items():
                    if k == ScheduledExam.EXAM_TIME_SCHEDULED:
                        total_exams_created_with_future_scheduling += len(v)
                        users_who_created_exams_with_future_scheduling.union(set(v))
                    elif k == ScheduledExam.PATIENT_CANCELED:
                        total_exams_created_with_cancelation += len(v)
                        users_who_created_exams_with_cancelation.union(set(v))
                    elif k == ScheduledExam.PROCEDURES_EXECUTED:
                        total_exams_created_attended_on_future += len(v)
                        users_who_created_exams_attended_on_future.union(set(v))

            todays_attended_exams = 0
            users_who_attended_exam = 0
            todays_scheduled_exams = 0
            users_who_scheduled_exam = 0
            users_with_scheduled_exams_for_today = set()
            users_with_calls_for_today = set()
            users_with_scheduled_exams_for_future = set()

            # todays_calls = ScheduledExamPhoneCall.objects.filter(call_time__date=date)
            # for call in todays_calls:
            #     users_with_calls_for_today.add(call.scheduled_exam.prescription.patient.user.email)

            exam = all_exams.get(date, None)
            # patient = exam.prescription.patient.user.email
            if exam:
                for k,v in exam.items():
                    if k == ScheduledExam.PROCEDURES_EXECUTED:
                        todays_attended_exams = v
                        users_who_attended_exam += int(v)
                    elif k == ScheduledExam.EXAM_TIME_SCHEDULED:
                        todays_scheduled_exams = v
                        users_who_scheduled_exam += int(v)



                    #     if status["value"] in [ScheduledExam.EXAM_TIME_SCHEDULED, ScheduledExam.LAB_RECORD_OPEN] \
                    #             and exam.scheduled_time and exam.scheduled_time.date() > date:
                    #         users_with_scheduled_exams_for_future.add(patient)
                    #
                    # if exam.scheduled_time and exam.scheduled_time.date() == date:
                    #     users_with_scheduled_exams_for_today.add(patient)

            row = (date, registrations.get(date, 0),
                   total_prescriptions_created, len(users_who_created_prescriptions),
                   total_calls_created, users_who_created_calls,
                   total_exams_created, len(users_who_created_exams),
                   total_exams_created_with_future_scheduling, len(users_who_created_exams_with_future_scheduling),
                   total_exams_created_with_cancelation, len(users_who_created_exams_with_cancelation),
                   total_exams_created_attended_on_future, len(users_who_created_exams_attended_on_future),
                   todays_attended_exams, users_who_attended_exam,
                   todays_scheduled_exams, users_who_scheduled_exam,
                   len(users_with_scheduled_exams_for_future)
                   )

            writer1.append(row)
            wb.save(file_name)
            if users_who_created_prescriptions:
                writer2.append((date,))
                for user in users_who_created_prescriptions:
                    writer2.append(('', user))

            # if users_who_created_calls:
            #     writer3.append((date,))
            #     for user in users_who_created_calls:
            #         writer3.append(('', user))

            if users_who_created_exams:
                writer4.append((date,))
                for user in users_who_created_exams:
                    writer4.append(('', user))

            if type(users_who_attended_exam) == list:
                writer5.append((date,))
                for user in users_who_attended_exam:
                    writer5.append(('', user))

            if type(users_who_scheduled_exam) == list:
                writer6.append((date,))
                for user in users_who_scheduled_exam:
                    writer6.append(('', user))

            if users_with_scheduled_exams_for_today:
                writer7.append((date,))
                for user in users_with_scheduled_exams_for_today:
                    writer7.append(('', user))

            if users_with_calls_for_today:
                writer8.append((date,))
                for user in users_with_calls_for_today:
                    writer8.append(('', user))

            if users_with_scheduled_exams_for_future:
                writer9.append((date,))
                for user in users_with_scheduled_exams_for_future:
                    writer9.append(('', user))

            # Sends file via email:
        text = """
                <html>
                    Hi!
                    <br><br>
                    This is your daily KPIs report since 15th June, 2017.
                    <br><br>
                    Sara
                    <br><br>
                    --
                    <br><br>
                    This is an automatic message, do not reply.
                </html>
                """
        cc = settings.USERS_REPORT_LIST + settings.DATA_REPORT_LIST
        response = Mail.send(to=cc,
                             subject="Daily KPIs Report",
                             text=text,
                             html=True,
                             attachments=[file_name, ])
        print("Email sent: {0}".format(response))

    except Exception as e:
        print(e)
        log.error(
            "Unable to build KPIs report {}"
                .format(e))



@app.task(queue="domain.tasks.send_operational_kpis")
def send_operational_kpis_report():
    """
    Sends operationsl kpis report daily
    :return:
    """
    from .utils import get_register_data

    # Restricting access:
    if settings.APP_ENVIRONMENT not in (settings.LOCAL, settings.PROD):
        log.warning("Restricting access to local and production environments only")
        return

    try:
        general_month = get_register_data('month')
        general_day = get_register_data('day')

        from domain.models import Patient, ScheduledExam, ScheduledExamPhoneCall
        from reporting.models import GeneralStatus
        from patient_api.utils import get_scheduled_exam_version_data, get_prescription_version

        PHONE_STATUS = (ScheduledExam.PHONE_CALL_NOT_ANSWERED,
                        ScheduledExam.PATIENT_CANCELED_BY_CALL,
                        ScheduledExam.EXAM_TIME_SCHEDULED)

        base = datetime.date(year=2017, month=7, day=1)
        today = datetime.date.today()
        numdays = (today - base).days

        phone_calls_query = ScheduledExamPhoneCall.objects.filter(created__gt=base)
        phone_calls_keys = []

        phone_calls = {'day':{},
                       'month': {}}

        for pc in phone_calls_query:
            if pc.phone not in phone_calls_keys:
                phone_calls_keys.append(pc.phone)
                ps = GeneralStatus.objects.filter(status_in=PHONE_STATUS,
                                                  exam=pc.scheduled_exam).order_by('date_set').first()

                if ps:
                    if ps.date_set >= pc.call_time:
                        tseconds = (ps.date_set - pc.call_time).totalseconds()
                        phone_calls['day'][pc.call_time.strftime('%Y-%m-%d')] = (tseconds, tseconds/86400)
                        phone_calls['month'][pc.call_time.strftime('%Y-%m')] = (tseconds, tseconds/86400)
                    else:
                        log.warning('Phone calls time doesnt match ScheduledExam ID {}'.format(pc.scheduled_exam))
                else:
                    ps = ScheduledExam.objects.filter(status_in=PHONE_STATUS,
                                                      id=pc.scheduled_exam)
                    if ps:
                        tseconds = (ps.modified - pc.call_time).totalseconds()
                        phone_calls['day'][pc.call_time.strftime('%Y-%m-%d')] = (tseconds, tseconds/86400)
                        phone_calls['month'][pc.call_time.strftime('%Y-%m')] = (tseconds, tseconds/86400)
                    else:
                        log.warning('Phone call not found ScheduledExam ID {}'.format(pc.scheduled_exam))


        #TODO: do the reporting XLS
        # tempo MÉDIO de resposta para prescrição - GERAL
        # tempo MÉDIO de resposta para prescrição - NOS TURNOS
        # tempo MÉDIO de resposta da ligação - entre horário agendado e ligação concluída

        date_list = [today - datetime.timedelta(days=x) for x in range(0, numdays+1)]

        sheet_headers = ("date", "time", "weekday",
                         "type", "seconds")

        file_name = "operational_kpis_per_day_report.xlsx"
        kpis_per_day_report_tab = "App"

        # Writes file:
        import os
        from openpyxl import Workbook
        wb = Workbook()
        writer1 = wb.active
        writer1.title = kpis_per_day_report_tab
        writer1.append(sheet_headers)
        writer2 = wb.create_sheet('Phone Schedule')
        writer2.append(sheet_headers)

        # Rows
        for date in date_list:

            # exams = ScheduledExam.objects.filter(created__contains=date)
            prescriptions = MedicalPrescription.objects.filter(created__contains=date)

            for prescription in prescriptions:
                start_date = prescription.created.strftime('%Y-%m-%d')
                start_time = prescription.created.strftime('%H:%M:%S')
                start_week = prescription.created.strftime('%w')

                prescription_statuses = get_scheduled_exam_version_data(prescription, ["status", "modified"], "OPERATION_STATUS", False)
                if prescription_statuses:
                    delta = datetime.datetime.strptime(prescription_statuses[0].get('modified'), "%Y-%m-%dT%H:%M:%S.%f") - prescription.created
                    writer1.append((start_date,
                                    start_time,
                                    start_week,
                                    prescription_statuses[0].get('status'),
                                    delta.total_seconds())
                                   )

                else:
                    try:
                        exam = prescription.scheduledexam_set.all()[0]
                        if exam:
                            exam_statuses = get_scheduled_exam_version_data(exam, ["status", "modified"], None, False)
                            if exam_statuses:
                                delta = datetime.datetime.strptime(exam_statuses[0].get('modified'), "%Y-%m-%dT%H:%M:%S.%f") - prescription.created
                                writer1.append((start_date,
                                                start_time,
                                                start_week,
                                                exam_statuses[0].get('status'),
                                                delta.total_seconds())
                                               )
                    except Exception as e:
                        pass

            for prescription in prescriptions:
                try:
                    exam = prescription.scheduledexam_set.all()[0]
                    if exam:
                        exam_statuses = get_scheduled_exam_version_data(exam, ["modified", "status"], None, True)
                        if exam_statuses:
                            i = 0
                            for k, v in exam_statuses.items():
                                if v in PHONE_STATUS:
                                    if len(exam_statuses) == 1:
                                        delta = prescription.created - datetime.datetime.strptime(k, "%Y-%m-%dT%H:%M:%S.%f")
                                        start_date = prescription.created.strftime('%Y-%m-%d')
                                        start_time = prescription.created.strftime('%H:%M:%S')
                                        start_week = prescription.created.strftime('%w')
                                    else:
                                        dstart = datetime.datetime.strptime(k, "%Y-%m-%dT%H:%M:%S.%f")
                                        delta = dstart - datetime.datetime.strptime(list(exam_statuses.keys())[i-1], "%Y-%m-%dT%H:%M:%S.%f")
                                        start_date = dstart.strftime('%Y-%m-%d')
                                        start_time = dstart.strftime('%H:%M:%S')
                                        start_week = dstart.strftime('%w')
                                    writer2.append((start_date,
                                                    start_time,
                                                    start_week,
                                                    v,
                                                    delta.total_seconds())
                                           )
                                i += 1
                except Exception as e:
                    pass

            wb.save(file_name)

            # Sends file via email:
        text = """
                <html>
                    Hi!
                    <br><br>
                    This is your daily KPIs report since 15th June, 2017.
                    <br><br>
                    Sara
                    <br><br>
                    --
                    <br><br>
                    This is an automatic message, do not reply.
                </html>
                """
        cc = settings.USERS_REPORT_LIST + settings.DATA_REPORT_LIST
        response = Mail.send(to=settings.USERS_REPORT_LIST,
                             subject="Daily Operational KPIs Report month day",
                             text=text,
                             html=True,
                             attachments=[file_name, ])
        print("Email sent: {0}".format(response))

    except Exception as e:
        print(e)
        log.error(
            "Unable to build KPIs report {}"
            .format(e))


@app.task(bind=True, queue="domain.tasks.create_images")
def create_images(
        self,
        get_pictures_method,
        patient_id=None,
        picture_id_front_uploadcare=None,
        picture_id_back_uploadcare=None,
        selfie_uploadcare=None,
        prescription_id=None,
        picture_insurance_card_front_uploadcare=None,
        picture_insurance_card_back_uploadcare=None ,
        picture_prescription_uploadcare=None,
        latest_prescription_pk=None,
        pieces=None
):
    """
    Uploads images to DEFAULT_FILE_STORAGE asynchronously
    :param self:
    :param get_pictures_method:
    :param patient_id:
    :param picture_id_front_uploadcare:
    :param picture_id_back_uploadcare:
    :param selfie_uploadcare:
    :param prescription_id:
    :param picture_insurance_card_front_uploadcare:
    :param picture_insurance_card_back_uploadcare:
    :param picture_prescription_uploadcare:
    :param latest_prescription_pk:
    :param pieces:
    :return:
    """

    import patient_api.utils as utils
    if patient_id:
        patient = Patient.objects.get(pk=patient_id)
        patient.picture_id_front_uploadcare = getattr(utils, get_pictures_method)(
            picture_id_front_uploadcare,
            "picture_id_front_uploadcare",
            patient,
            latest_prescription_pk=latest_prescription_pk
        )
        patient.picture_id_back_uploadcare = getattr(utils, get_pictures_method)(
            picture_id_back_uploadcare,
            "picture_id_back_uploadcare",
            patient,
            latest_prescription_pk=latest_prescription_pk
        )
        patient.selfie_uploadcare = getattr(utils, get_pictures_method)(
            selfie_uploadcare,
            "selfie_uploadcare",
            patient,
            latest_prescription_pk=latest_prescription_pk
        )
        patient.save()

    if prescription_id:
        prescription = MedicalPrescription.objects.get(pk=prescription_id)

        prescription.picture_insurance_card_front_uploadcare = getattr(utils, get_pictures_method)(
            picture_insurance_card_front_uploadcare,
            "picture_insurance_card_front_uploadcare",
            latest_prescription_pk=latest_prescription_pk
        )
        prescription.picture_insurance_card_back_uploadcare = getattr(utils, get_pictures_method)(
            picture_insurance_card_back_uploadcare,
            "picture_insurance_card_back_uploadcare",
            latest_prescription_pk=latest_prescription_pk
        )
        prescription.picture_prescription_uploadcare = getattr(utils, get_pictures_method)(
            picture_prescription_uploadcare,
            "picture_prescription_uploadcare",
            latest_prescription_pk=latest_prescription_pk
        )

        prescription.picture_id_front_uploadcare = prescription.patient.picture_id_front_uploadcare
        prescription.picture_id_back_uploadcare = prescription.patient.picture_id_back_uploadcare
        prescription.selfie_uploadcare = prescription.patient.selfie_uploadcare
        prescription.save()

        if pieces:
            for piece_data in pieces:
                piece = PrescriptionPiece.objects.get(pk=piece_data.get("pk"))
                prescription_image = piece_data.get("picture", None)
                if prescription_image:
                    piece.picture = utils.get_piece_picture_as_content_file_uploadcare(
                        piece_data.get("picture"),
                        "picture_prescription_uploadcare",
                    )
                    piece.save()
        else:
            prescription_peaces = PrescriptionPiece.objects.filter(prescription=prescription)
            if not prescription_peaces.exists():
                prescription_peace = PrescriptionPiece(prescription=prescription,
                                                       picture=prescription.picture_prescription_uploadcare)
                prescription_peace.save()

            if len(prescription_peaces) == 1:
                prescription_peace = prescription_peaces.first()
                prescription_peace.picture = prescription.picture_prescription_uploadcare
                prescription_peace.save()

        # Updates Concierge Firebase only after all async image uploading process is done for the entire prescription:
        if prescription.status == MedicalPrescription.PATIENT_REQUESTED:
            from domain.signals.update_firebase_db import sync_prescription_to_firebase_after_async_upload
            sync_prescription_to_firebase_after_async_upload(prescription)
            try:
                if settings.APP_ENVIRONMENT in (settings.PROD, settings.STAGING):
                    zendesk_prescription.apply_async(args=[prescription.id], countdown=5)
            except Exception as e:
                log.error("Unable to create zendesk ticket.\n {}".format(e))

@app.task(bind=True, queue="domain.tasks.update_results")
def update_results(self, data=None):
    """
    process the received results from Sara Clinic
    :param self:
    :param data:
    :return:
    """

    if settings.APP_ENVIRONMENT not in (settings.PROD, settings.STAGING, settings.LOCAL):
        log.warning("Restricting firebase access to local, staging and production environments only")
        return

    for item in data:
        sexams = ScheduledExam.objects.filter(lab_file_code=str(item), results_arrived_at=None)
        for sexam in sexams:
            try:
                sexam.results_arrived_at = datetime.datetime.now()
                sexam.status = ScheduledExam.RESULTS_RECEIVED
                sexam.save()
            except Exception as e:
                log.warning("Could not mark the result exam for Lab file code: {} and Schedule Exam: {}.\n {}".format(item, sexam.pk, e))
